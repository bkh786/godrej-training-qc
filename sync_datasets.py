#!/usr/bin/env python3
"""
Automated Data Synchronizer for Godrej Training & Merchandising Dashboards
Downloads latest live datasets directly from SharePoint public export URLs,
embeds them into index.html, training.html, and m_score.html, and automatically
pushes changes to GitHub.
"""

import urllib.request
import http.cookiejar
from pyxlsb import open_workbook
import openpyxl
import io
import json
import datetime
import os
import subprocess
import sys

SHAREPOINT_URLS = {
    "index": "https://teamchannelplay-my.sharepoint.com/:x:/g/personal/bikash_roy1_channelplay_in/IQBx5HIst0LPT4_moEtMpsbtAd4w3ClOl0h-mrlnCEmDCno?download=1",
    "training": "https://teamchannelplay-my.sharepoint.com/:x:/g/personal/bikash_roy1_channelplay_in/IQBRmCEH6nI8TLFuK-RVqPu0ATXbidF7rGfITZvpZH7PyAA?download=1",
    "m_score": "https://teamchannelplay-my.sharepoint.com/:x:/g/personal/bikash_roy1_channelplay_in/IQAaW2sHEFKnRrqPtppHxBH2ARgiE7222JHi46SCAbbXkQ8?download=1",
    "program_performance": "https://teamchannelplay-my.sharepoint.com/:x:/g/personal/bikash_roy1_channelplay_in/IQB-EkkYdgTFQphMhXNUEfKSAWIldx1iKI_TWThpQI42w8E?e=2dQVkl&download=1"
}

def get_opener():
    cookie_jar = http.cookiejar.CookieJar()
    return urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cookie_jar),
        urllib.request.HTTPRedirectHandler
    )

def embed_sample_data(html_file, rows):
    if not os.path.exists(html_file):
        print(f"File {html_file} not found, skipping.")
        return
    with open(html_file, "r", encoding="utf-8") as f:
        html = f.read()

    compact_json = json.dumps(rows, separators=(",", ":"))
    tag_start = '<script id="sample-data" type="application/json">'
    tag_end = '</script>'
    
    idx1 = html.find(tag_start)
    if idx1 != -1:
        idx2 = html.find(tag_end, idx1)
        if idx2 != -1:
            html = html[:idx1 + len(tag_start)] + compact_json + html[idx2:]
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"✓ Successfully embedded {len(rows):,} rows into {html_file}")

def sync_index(opener):
    print("\n--- Syncing index.html (QC Tracker) ---")
    req = urllib.request.Request(SHAREPOINT_URLS["index"], headers={"User-Agent": "Mozilla/5.0"})
    data = opener.open(req, timeout=30).read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "qc" in s.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        if any(v is not None for v in r):
            row_clean = []
            for cell in r:
                if isinstance(cell, (datetime.date, datetime.datetime)):
                    row_clean.append(cell.strftime("%Y-%m-%d"))
                else:
                    row_clean.append(cell)
            rows.append(row_clean)
    embed_sample_data("index.html", rows)

def sync_training(opener):
    print("\n--- Syncing training.html (Training Details) ---")
    req = urllib.request.Request(SHAREPOINT_URLS["training"], headers={"User-Agent": "Mozilla/5.0"})
    data = opener.open(req, timeout=30).read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "training" in s.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        if any(v is not None for v in r):
            row_clean = []
            for cell in r:
                if isinstance(cell, (datetime.date, datetime.datetime)):
                    row_clean.append(cell.strftime("%Y-%m-%d"))
                else:
                    row_clean.append(cell)
            rows.append(row_clean)
    embed_sample_data("training.html", rows)

def sync_m_score(opener):
    print("\n--- Syncing m_score.html (Product-VM-Score) ---")
    req = urllib.request.Request(SHAREPOINT_URLS["m_score"], headers={"User-Agent": "Mozilla/5.0"})
    data = opener.open(req, timeout=45).read()
    with open_workbook(io.BytesIO(data)) as wb:
        sheet_name = "Product-VM-Score" if "Product-VM-Score" in wb.sheets else wb.sheets[0]
        with wb.get_sheet(sheet_name) as s:
            rows = []
            for i, row in enumerate(s.rows()):
                r_vals = [c.v for c in row[:30]]
                if i > 0 and len(r_vals) > 14:
                    d = r_vals[14]
                    if isinstance(d, (int, float)):
                        dt = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(d))
                        r_vals[14] = dt.strftime("%Y-%m-%d")
                rows.append(r_vals)
            embed_sample_data("m_score.html", rows)

def sync_program_performance(opener):
    print("\n--- Syncing program_performance.html (Program Performance) ---")
    req = urllib.request.Request(SHAREPOINT_URLS["program_performance"], headers={"User-Agent": "Mozilla/5.0"})
    data = opener.open(req, timeout=45).read()
    with open_workbook(io.BytesIO(data)) as wb:
        sheet_name = "Program Performance" if "Program Performance" in wb.sheets else wb.sheets[0]
        with wb.get_sheet(sheet_name) as s:
            rows = []
            for i, row in enumerate(s.rows()):
                r_vals = [c.v for c in row]
                # convert Excel serial dates
                if i == 0 and len(r_vals) > 5 and isinstance(r_vals[5], (int, float)) and r_vals[5] > 20000:
                    r_vals[5] = (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(r_vals[5]))).strftime("%Y-%m-%d")
                if i >= 2 and len(r_vals) > 0 and isinstance(r_vals[0], (int, float)) and r_vals[0] > 20000:
                    r_vals[0] = (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(r_vals[0]))).strftime("%Y-%m-%d")
                rows.append(r_vals)
            embed_sample_data("program_performance.html", rows)

def auto_push_to_github():
    print("\n--- Pushing updates to GitHub ---")
    try:
        subprocess.run(["git", "add", "index.html", "training.html", "m_score.html", "program_performance.html"], check=True)
        # Check if there are changes to commit
        res = subprocess.run(["git", "diff", "--staged", "--quiet"])
        if res.returncode != 0:
            msg = f"chore(data): auto-sync latest live SharePoint datasets ({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})"
            subprocess.run(["git", "commit", "-m", msg], check=True)
            subprocess.run(["git", "push", "origin", "main"], check=True)
            print("🚀 Successfully pushed updated datasets to GitHub main branch!")
        else:
            print("No data changes detected. Remote repository is already up-to-date.")
    except Exception as e:
        print("Note: Could not push to git automatically:", e)

def main():
    print("Starting SharePoint Data Sync...")
    opener = get_opener()
    sync_index(opener)
    sync_training(opener)
    sync_m_score(opener)
    sync_program_performance(opener)
    print("\nAll datasets synchronized successfully!")
    auto_push_to_github()

if __name__ == "__main__":
    main()
